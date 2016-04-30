# !/usr/bin/env python

'''\n  CSR generator

  Simple script to generate CSR's and private keys from a file or single entry on the command line.


  Entries should be in a form of fqdn/url or box name.

  Some defaults are pre-set for ease of use:
  such as  location = "Smallville"
    if this needs to change to "CHESAPEAKE"
    then  "-L CHESAPEAKE" would be used


  -f for processing a file type of either text (.txt) or excel (.xlsx)
  -n for node (single entry)

  The sub-directories for CSR's and keys will be created if not present.
  This is where the script will place generated CSR's and it's keys.



  Usage:
  python csr_generator.py  -f <file name>
  python csr_generator.py  -n node

  python csr_generator.py  -n or -f <?>    -L <?>  -S <?>  -O <?>  -U <?> "

  Arguments:
    -f      file - list of nodes

    -n      single node on command line

    -S      State
                default = "some state?"

    -L      Location of cert
              default = "some town?"

    -O      Organization Name
              default = "an org?"

    -U      Organization Unit Name
              default = "an org's parent"


  Flags:
    -h      help

    Version: 002

'''

# Libraries/Modules
from OpenSSL import crypto, SSL
import sys
import os
import getopt
import openpyxl
import re


def usage():
    """
    Describes general usage that includes options,
    arguments, lags
    """
    print(__doc__)


def pars_cmd(argv):
    """

    #sets a simple internel default to check file options
    # renamed some of the vars for ease of use.
    # in general this would be a naming convention.
    # Allows you to permanently set values required for CSR
    # C  = 'US'
    # ST = 'New York'
    # L  = 'Location'
    # O  = 'Organization'
    # OU = 'Organizational Unit'
    ###############################

    """

    cert_st = "some state?"
    cert_loc = "some town?"
    cert_org = "an org"
    cert_org_unit = "an orgs parent"

    file_name = None
    data_flag = 0

    if (len(sys.argv[1:]) != 0):

        try:
            opts, args = getopt.getopt(sys.argv[1:], "n:f:L:S:O:U:h")
        except getopt.GetoptError as err:
            # print help information and exit:
            print(err)  # will print something like "option -a not recognized"
            usage()
            sys.exit(2)

        # checks to see if opts(list) is empty

        if not opts:
            print("There were no options given use -h to see help\n")

        # parses opts and arguments
        for o, a in opts:

            if o in ("-h"):
                # help
                usage()
                exit()

            elif o in ("-f"):
                # multi-node file
                file_name = a
                data_flag = 1

            elif o in ("-n"):
                # single node on command line
                file_name = a
                data_flag = 2

            # cert requirments
            elif o in ("-S"):
                cert_st = a
                cert_st = str(cert_st)
            elif o in ("-L"):
                cert_loc = a
                cert_loc = str(cert_loc)
            elif o in ("-O"):
                cert_org = a
                cert_org = str(cert_org)
            elif o in ("-U"):
                cert_org_unit = a
                cert_org_unit = str(cert_org_unit)
            else:
                assert False, "parse - unhandled option"
                # ...
                # change options check as needed.

    return (data_flag, file_name, cert_st, cert_loc, cert_org, cert_org_unit)


def list_builder(data_flag, file_name):
    """

    # takes file name /data_flag to determine if it is from commmand line or a file
    # if it is a file, checks to see what type of file
    #
    # if it is a text file, then builds a many element dictionary

    # if it is an excel file, it will read over the rows/collums at pre-set area
    #  pre-set dependencies here ----
    #          fqdn is in collum "B" = cell_B_value
    #          secondary names are in collum "C" = cell_C_value

    # if a string, dictionary will contain single entry.
    """

    name_list = {}

    # collum description and name
    c_fqdn = 'B'
    c_alt_name = 'C'
    c_disa = 'H'

    # testing to see if data flag is for a file
    if data_flag == 1:
        #
        # testing for file types text
        if file_name.endswith('.txt'):
            try:
                with open(file_name) as file:
                    for line_data in file:
                        name = line_data.strip()
                        if len(line_data) > 1:
                            # a "1" is putin for key values
                            # as a placeholder for future use
                            name_list[name] = 1
            except IOError as e:
                print(" Unable to open text file")

        # testing to see if it is an excle
        elif file_name.endswith('xlsx'):
            try:
                wb = openpyxl.load_workbook(file_name, read_only=True)
                ws = wb.active

                # spread sheet data starts on row 2.
                incr = 1
                cell_B_value = ws[c_fqdn + str(incr)].value

                # iterate through B collum data.
                while cell_B_value != None:

                    incr += 1
                    cell_B_value = ws[c_fqdn + str(incr)].value
                    cell_H_value = str(ws[c_disa + str(incr)].value)

                    # building dictonary
                    #
                    if cell_B_value is not None:

                        if cell_H_value == "None":
                            cell_C_value = str(ws[c_alt_name + str(incr)].value)

                            # building value into key
                            if cell_C_value == "None":
                                name_list[cell_B_value] = 1
                            else:
                                name_list[cell_B_value] = cell_C_value

            except IOError as e:
                print(" Unable to open excel file")
        else:
            usage()
            exit()
    #
    # testing to see if data flag is for command line
    elif data_flag == 2:
        name = file_name.strip()
        name_list[name] = 1
    else:
        assert False, "file - unhandled option"

    return (name_list)


def generateKey(type, bits):
    """
    # generates a single instance of a private key
    # for each csr generation

    """

    key = crypto.PKey()
    key.generate_key(type, bits)
    return key


def generateFiles(flag, nodename, request):
    """
    # generates csr using ...
    #
    """

    if flag == 1:
        f = open(os.path.join('csr', nodename), "wb")
        # print("dump :" + str(crypto.dump_certificate_request(crypto.FILETYPE_PEM, request)))
        f.write(crypto.dump_certificate_request(crypto.FILETYPE_PEM, request))
        f.close()

    elif flag == 2:
        f = open(os.path.join('keys', nodename), "wb")
        f.write(crypto.dump_privatekey(crypto.FILETYPE_PEM, request))
        f.close()
    else:
        print("Failed.")
        exit()


def generateCSR(nodename, cert_st, cert_loc, cert_org, cert_org_unit):
    """
    # generates csr using ...
    #
    """
    csrfile = nodename + '.csr'
    keyfile = nodename + '.key'
    TYPE_RSA = crypto.TYPE_RSA

    req = crypto.X509Req()
    req.get_subject().CN = nodename
    req.get_subject().countryName = "US"
    req.get_subject().stateOrProvinceName = cert_st
    req.get_subject().localityName = cert_loc
    req.get_subject().organizationName = cert_org
    req.get_subject().organizationalUnitName = cert_org_unit

    # Utilizes generateKey function to kick off key generation.
    key = generateKey(TYPE_RSA, 2048)
    req.set_pubkey(key)

    # hash/finger print is set to 256 and not "1"
    req.sign(key, "sha256")

    generateFiles(1, csrfile, req)
    generateFiles(2, keyfile, key)

    # later use of a flag
    return req


def generateCSR_wtih_SAN(nodename, sans, cert_st, cert_loc, cert_org, cert_org_unit):
    """
    # generates csr using ...
    #
    """

    csrfile = nodename + '.csr'
    keyfile = nodename + '.key'
    TYPE_RSA = crypto.TYPE_RSA

    req = crypto.X509Req()
    req.get_subject().CN = nodename
    req.get_subject().countryName = "US"
    req.get_subject().stateOrProvinceName = cert_st
    req.get_subject().localityName = cert_loc
    req.get_subject().organizationName = cert_org
    req.get_subject().organizationalUnitName = cert_org_unit

    # container for x509 extentions ie altnames list
    extentions = []

    if sans:
        sans = sans.split(",")

        altname = b'subjectAltName'
        dns = b'DNS:'

        for node in sans:
            node = node.strip()
            dns_name = dns + node.encode()
            extentions.append(crypto.X509Extension(altname, False, dns_name))

    req.add_extensions(extentions)

    # Utilizes generateKey function to kick off key generation.
    key = generateKey(TYPE_RSA, 2048)
    req.set_pubkey(key)

    # hash/finger print is set to 256 and not "1"
    req.sign(key, "sha256")

    generateFiles(1, csrfile, req)
    generateFiles(2, keyfile, key)

    # later use of a flag
    return req


def main():
    file_name = None

    data_flag, file_name, cert_st, cert_loc, cert_org, cert_org_unit = pars_cmd(sys.argv)

    if data_flag == 0:
        usage()
        exit()

    # check to see if sub dir's are in place if not make
    # two below execution point
    d_csr = 'csr'
    d_key = 'keys'

    if not os.path.exists(d_csr):
        os.makedirs(d_csr)
    if not os.path.exists(d_key):
        os.makedirs(d_key)

    list_of_names = list_builder(data_flag, file_name)

    # processing dictionary
    for d_nodename in list_of_names:
        sans = []
        d_nodename = d_nodename.rstrip()
        if len(d_nodename) > 0:
            # set consistency of nodename to lower case

            if list_of_names[d_nodename] == 1:
                d_nodename = d_nodename.lower()
                generateCSR(d_nodename, cert_st, cert_loc, cert_org, cert_org_unit)
            else:
                sans = list_of_names[d_nodename]
                d_nodename = d_nodename.lower()
                generateCSR_wtih_SAN(d_nodename, sans, cert_st, cert_loc, cert_org, cert_org_unit)

    print("\n\n  Processing is complete\n  CSR's will be in a subdirectory called \"CSR\" ")


if __name__ == "__main__":
    main()
